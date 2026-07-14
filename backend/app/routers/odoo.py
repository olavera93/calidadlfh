import io
from datetime import datetime
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import Response
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.dependencies import get_current_user, require_admin
from app.crud import odoo as crud_odoo
from app.database import get_db
from app.models.models import Usuario
from app.schemas.odoo import (
    OdooSettingsCreate,
    OdooSettingsOut,
    OdooTestResult,
    RecepcionItem,
    VencimientoItem,
)
from app.services import odoo_service as svc

router = APIRouter(prefix="/odoo", tags=["odoo"])

HDR_FILL = PatternFill("solid", fgColor="1E293B")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)


def _get_credentials(db: Session):
    setting = crud_odoo.get_settings(db)
    if not setting or not setting.url:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Odoo no está configurado. Configure las credenciales en Ajustes.",
        )
    return setting


# ── Configuración ──────────────────────────────────────────────────────────────

@router.get("/settings", response_model=OdooSettingsOut)
def get_settings(
    db: Session = Depends(get_db),
    _: Usuario = Depends(require_admin),
):
    setting = crud_odoo.get_settings(db)
    if not setting:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                            detail="No hay configuración de Odoo guardada.")
    return setting


@router.put("/settings", response_model=OdooSettingsOut)
def save_settings(
    data: OdooSettingsCreate,
    db: Session = Depends(get_db),
    _: Usuario = Depends(require_admin),
):
    try:
        return crud_odoo.upsert_settings(db, data)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/settings/test", response_model=OdooTestResult)
def test_connection(
    db: Session = Depends(get_db),
    _: Usuario = Depends(require_admin),
):
    setting = crud_odoo.get_settings(db)
    if not setting or not setting.url:
        return OdooTestResult(ok=False, mensaje="No hay configuración guardada.")
    try:
        uid = svc.authenticate(setting.url, setting.database,
                               setting.username, setting.password)
        crud_odoo.update_uid(db, setting, uid)
        return OdooTestResult(ok=True, mensaje="Conexión exitosa.", uid=uid)
    except svc.OdooConnectionError as e:
        return OdooTestResult(ok=False, mensaje=str(e))


# ── Recepciones ────────────────────────────────────────────────────────────────

@router.get("/recepciones", response_model=list[RecepcionItem])
def get_recepciones(
    fecha_inicio:      Optional[str] = Query(None),
    fecha_fin:         Optional[str] = Query(None),
    ubicacion_origen:  Optional[str] = Query(None),
    ubicacion_destino: Optional[str] = Query(None),
    numero_movimiento: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    setting = _get_credentials(db)
    try:
        return svc.get_recepciones(
            setting.url, setting.database, setting.uid, setting.password,
            fecha_inicio, fecha_fin, ubicacion_origen, ubicacion_destino, numero_movimiento,
        )
    except svc.OdooConnectionError as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))


@router.get("/recepciones/xlsx")
def get_recepciones_xlsx(
    fecha_inicio:      Optional[str] = Query(None),
    fecha_fin:         Optional[str] = Query(None),
    ubicacion_origen:  Optional[str] = Query(None),
    ubicacion_destino: Optional[str] = Query(None),
    numero_movimiento: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    setting = _get_credentials(db)
    try:
        data = svc.get_recepciones(
            setting.url, setting.database, setting.uid, setting.password,
            fecha_inicio, fecha_fin, ubicacion_origen, ubicacion_destino, numero_movimiento,
        )
    except svc.OdooConnectionError as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))

    headers    = ["Fecha", "Origen", "Destino", "Proveedor", "N° Orden",
                  "Factura", "Producto", "Lote", "Vencimiento", "Cantidad"]
    col_widths = [12, 20, 20, 30, 15, 18, 40, 18, 14, 10]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recepciones"
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(vertical="center")
    for ri, r in enumerate(data, 2):
        orden_txt = r.orden + (" (inferida)" if r.orden_inferida else "")
        ws.cell(row=ri, column=1,  value=r.fecha)
        ws.cell(row=ri, column=2,  value=r.origen)
        ws.cell(row=ri, column=3,  value=r.destino)
        ws.cell(row=ri, column=4,  value=r.proveedor)
        ws.cell(row=ri, column=5,  value=orden_txt)
        ws.cell(row=ri, column=6,  value=r.factura)
        ws.cell(row=ri, column=7,  value=r.producto)
        ws.cell(row=ri, column=8,  value=r.lote)
        ws.cell(row=ri, column=9,  value=r.vencimiento)
        ws.cell(row=ri, column=10, value=r.cantidad)
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    fecha = datetime.now().strftime("%Y-%m-%d")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=recepciones_{fecha}.xlsx"},
    )


# ── Vencimientos ───────────────────────────────────────────────────────────────

@router.get("/vencimientos", response_model=list[VencimientoItem])
def get_vencimientos(
    fecha_hasta: str = Query(...),
    fecha_desde: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    setting = _get_credentials(db)
    try:
        return svc.get_vencimientos(
            setting.url, setting.database, setting.uid, setting.password,
            fecha_hasta, fecha_desde,
        )
    except svc.OdooConnectionError as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))


@router.get("/vencimientos/xlsx")
def get_vencimientos_xlsx(
    fecha_hasta: str = Query(...),
    fecha_desde: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    setting = _get_credentials(db)
    try:
        data = svc.get_vencimientos(
            setting.url, setting.database, setting.uid, setting.password,
            fecha_hasta, fecha_desde,
        )
    except svc.OdooConnectionError as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))

    headers    = ["Vencimiento", "Días restantes", "Producto", "Lote", "Ubicación", "Cantidad"]
    col_widths = [14, 14, 45, 20, 30, 12]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vencimientos"
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(vertical="center")
    for ri, r in enumerate(data, 2):
        dias  = r.dias
        color = ("DC2626" if dias is not None and dias < 30
                 else "D97706" if dias is not None and dias < 90
                 else "16A34A")
        ws.cell(row=ri, column=1, value=r.vencimiento)
        c = ws.cell(row=ri, column=2, value=dias)
        c.font = Font(color=color, bold=True)
        ws.cell(row=ri, column=3, value=r.producto)
        ws.cell(row=ri, column=4, value=r.lote)
        ws.cell(row=ri, column=5, value=r.ubicacion)
        ws.cell(row=ri, column=6, value=r.cantidad)
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    fecha = datetime.now().strftime("%Y-%m-%d")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=vencimientos_{fecha}.xlsx"},
    )
