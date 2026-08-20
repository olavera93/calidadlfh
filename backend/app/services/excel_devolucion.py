import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def generar_excel_devolucion(devoluciones: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "FORMATO DEVOLUCIONES"
    ws.views.sheetView[0].showGridLines = True

    # Estilos
    font_bold = Font(name="Arial", size=9, bold=True)
    font_regular = Font(name="Arial", size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Anchos de columna (A a P)
    col_widths = {
        "A": 14,  # N° FORMATO
        "B": 35,  # DESCRIPCIÓN PRODUCTO
        "C": 18,  # FORMA FARMACÉUTICA
        "D": 12,  # LOTE
        "E": 14,  # FECHA VENCIMIENTO
        "F": 18,  # REGISTRO SANITARIO
        "G": 20,  # LABORATORIO
        "H": 20,  # PROVEEDOR
        "I": 10,  # CANTIDAD
        "J": 25,  # CAUSA DE DEVOLUCIÓN
        "K": 14,  # FECHA CREACIÓN
        "L": 14,  # FECHA RECIBIDO
        "M": 14,  # FECHA ENTREGADO
        "N": 22,  # QUIEN RECIBE
        "O": 22,  # QUIEN ENTREGA
        "P": 15,  # ESTADO
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    primer_item = devoluciones[0] if devoluciones else None
    
    # Fecha de creación para la cabecera única
    fecha_creacion_gen = ""
    if primer_item and getattr(primer_item, "fecha_creacion", None):
        fecha_creacion_gen = str(primer_item.fecha_creacion.date()) if hasattr(primer_item.fecha_creacion, "date") else str(primer_item.fecha_creacion)

    # Encabezado superior simplificado (solo la fecha)
    ws["A1"] = f"FECHA CREACIÓN: {fecha_creacion_gen}"
    ws["A1"].font = font_bold

    # Encabezados de la grilla principal
    headers = [
        ("A3", "N° FORMATO"),
        ("B3", "DESCRIPCIÓN DEL PRODUCTO FARMACÉUTICO"),
        ("C3", "FORMA FARMACÉUTICA"),
        ("D3", "LOTE"),
        ("E3", "FECHA DE VENCIMIENTO"),
        ("F3", "REGISTRO SANITARIO"),
        ("G3", "LABORATORIO"),
        ("H3", "PROVEEDOR"),
        ("I3", "CANTIDAD"),
        ("J3", "CAUSA DE DEVOLUCIÓN"),
        ("K3", "FECHA CREACIÓN"),
        ("L3", "FECHA RECIBIDO"),
        ("M3", "FECHA ENTREGADO"),
        ("N3", "QUIEN RECIBE"),
        ("O3", "QUIEN ENTREGA"),
        ("P3", "ESTADO"),
    ]

    for cell_ref, text in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = font_bold
        cell.fill = gray_fill
        cell.alignment = center_align

    # Filas de datos
    start_row = 4
    
    for i, dev in enumerate(devoluciones):
        current_row = start_row + i

        num_formato_item = getattr(dev, "numero_de_formato", "") if dev else ""
        if not num_formato_item and dev and getattr(dev, "id", None):
            num_formato_item = str(dev.id).zfill(3)

        # Laboratorio desde la devolución o relación de producto
        lab_nombre = ""
        if dev:
            if getattr(dev, "laboratorio", None):
                lab_nombre = dev.laboratorio
            elif getattr(dev, "producto", None) and getattr(dev.producto, "laboratorio", None):
                lab_nombre = dev.producto.laboratorio

        # Formateo de fechas
        f_venc = str(dev.fecha_de_vencimiento) if dev and getattr(dev, "fecha_de_vencimiento", None) else ""
        f_creac = str(dev.fecha_creacion.date()) if dev and getattr(dev, "fecha_creacion", None) and hasattr(dev.fecha_creacion, "date") else str(getattr(dev, "fecha_creacion", ""))
        f_rec = str(dev.fecha_recibido) if dev and getattr(dev, "fecha_recibido", None) else ""
        f_ent = str(dev.fecha_entregado) if dev and getattr(dev, "fecha_entregado", None) else ""

        # Nombres de quien recibe y entrega
        q_recibe = dev.quien_recibe if dev and getattr(dev, "quien_recibe", None) else ""
        q_entrega = dev.quien_entrega if dev and getattr(dev, "quien_entrega", None) else ""
        
        # Estado
        estado_val = dev.estado if dev and getattr(dev, "estado", None) else "Pendiente"

        ws[f"A{current_row}"] = num_formato_item
        ws[f"B{current_row}"] = dev.producto.nombre if dev and getattr(dev, "producto", None) else ""
        ws[f"C{current_row}"] = dev.forma_farmaceutica if dev else ""
        ws[f"D{current_row}"] = dev.lote if dev else ""
        ws[f"E{current_row}"] = f_venc
        ws[f"F{current_row}"] = dev.registrosanitario if dev else ""
        ws[f"G{current_row}"] = lab_nombre
        ws[f"H{current_row}"] = dev.proveedor.nombre if dev and getattr(dev, "proveedor", None) else ""
        ws[f"I{current_row}"] = dev.cantidad if dev else ""
        ws[f"J{current_row}"] = dev.causa if dev else ""
        ws[f"K{current_row}"] = f_creac
        ws[f"L{current_row}"] = f_rec
        ws[f"M{current_row}"] = f_ent
        ws[f"N{current_row}"] = q_recibe
        ws[f"O{current_row}"] = q_entrega
        ws[f"P{current_row}"] = estado_val

        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
            c = ws[f"{col}{current_row}"]
            c.font = font_regular
            c.alignment = center_align if col not in ["B", "G", "H", "J", "N", "O"] else left_align

    # Aplicar bordes al encabezado A1 y a toda la tabla
    ws["A1"].border = thin_border
    
    end_row = start_row + len(devoluciones) - 1 if devoluciones else start_row
    for row in ws.iter_rows(min_row=3, max_row=end_row, min_col=1, max_col=16):
        for cell in row:
            cell.border = thin_border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output